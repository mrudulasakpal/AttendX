import os
import openpyxl
from datetime import datetime

EXCEL_FILE = "Attendance_Sheet.xlsx"

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Date", "Time", "Session ID", "Roll Number", "Status"])
        workbook.save(EXCEL_FILE)

def mark_attendance(roll_number: str, session_id: int, session_time: datetime):
    initialize_excel()
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    
    # Check if already marked for this specific session in Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2] == session_id and row[3] == roll_number:
            return False # already marked
            
    now = datetime.now()
    sheet.append([
        session_time.strftime("%Y-%m-%d"), 
        session_time.strftime("%H:%M:%S"), 
        session_id,
        roll_number, 
        "Present"
    ])
    workbook.save(EXCEL_FILE)
    return True
